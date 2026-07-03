#!/usr/bin/env python3
"""
Migrate outreach XLSX into Supabase organizations table.

Usage:
    python migrate_sheets.py Outreach.xlsx
    python migrate_sheets.py Outreach.xlsx --dry-run

Sheet layout:
    Row 1 — spreadsheet title   (skip)
    Row 2 — blank / meta        (skip)
    Row 3 — column headers
    Row 4+ — data

Requires environment variables:
    SUPABASE_URL         - e.g. https://xxxx.supabase.co
    SUPABASE_SERVICE_KEY - service role key (bypasses RLS)
    CHAPTER_ID           - chapter UUID (overridden by --chapter-id flag)
"""

import json
import os
import sys
import argparse
from datetime import datetime
import urllib.request
import urllib.error

BATCH_SIZE = 50

COLUMN_MAP = {
    'Organization Name':    'org_name',
    'Organization Type':    'org_type',
    'Website':              'website',
    'Contact Name':         'contact_name',
    'Position / Title':     'contact_title',
    'Email Address':        'email',
    'Phone Number':         'phone',
    'Township/ State':      'township',
    'Date Researched':      'date_researched',
    'Date First Contacted': 'date_first_contacted',
    'Contact Method':       'contact_method',
    'Current Status':       'current_status',
    'Follow-Up Date':       'follow_up_date',
    'Last Response Date':   'last_response_date',
    'Partnership Interest': 'partnership_interest',
    'Notes':                'notes',
    'Outcome':              'outcome',
    # logged_by → skip (UUID FK, left NULL)
}

DATE_FIELDS = {
    'date_researched', 'date_first_contacted',
    'follow_up_date', 'last_response_date',
}


def parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if not s or s == 'None':
        return None
    for fmt in ('%m/%d/%Y', '%m/%d/%y', '%Y-%m-%d'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return None


def cell_str(v):
    if v is None:
        return ''
    s = str(v).strip()
    return '' if s == 'None' else s


def map_row(xlsx_row, chapter_id):
    record = {'chapter_id': chapter_id}
    for xlsx_col, db_col in COLUMN_MAP.items():
        raw = xlsx_row.get(xlsx_col)
        if db_col in DATE_FIELDS:
            val = parse_date(raw)
        else:
            val = cell_str(raw) or None
        record[db_col] = val
    return record


def insert_batch(batch, url, headers, dry_run):
    if dry_run:
        for row in batch:
            print(f"  [dry-run] row {row.get('row_number')} — {row.get('org_name', '?')}")
        return 0

    body = json.dumps(batch).encode('utf-8')
    req = urllib.request.Request(
        f"{url}/rest/v1/organizations",
        data=body,
        headers={**headers, 'Content-Type': 'application/json', 'Prefer': 'return=minimal'},
        method='POST',
    )
    try:
        with urllib.request.urlopen(req) as resp:
            if resp.status not in (200, 201):
                raise RuntimeError(f"HTTP {resp.status}")
        return 0
    except urllib.error.HTTPError as e:
        body_text = e.read().decode('utf-8', errors='replace')
        print(f"  ERROR inserting batch: HTTP {e.code} — {body_text}", file=sys.stderr)
        return len(batch)


def main():
    parser = argparse.ArgumentParser(description='Migrate outreach XLSX to Supabase')
    parser.add_argument('xlsx_path', help='Path to the Outreach.xlsx file')
    parser.add_argument('--chapter-id',
                        default=os.environ.get('CHAPTER_ID', 'e554fa97-f949-4600-8023-b65d60edd034'),
                        help='Chapter UUID')
    parser.add_argument('--dry-run', action='store_true',
                        help='Print rows without inserting')
    args = parser.parse_args()

    supabase_url = os.environ.get('SUPABASE_URL', '').rstrip('/')
    service_key  = os.environ.get('SUPABASE_SERVICE_KEY', '')

    if not args.dry_run:
        if not supabase_url:
            print('ERROR: SUPABASE_URL env var not set', file=sys.stderr)
            sys.exit(1)
        if not service_key:
            print('ERROR: SUPABASE_SERVICE_KEY env var not set', file=sys.stderr)
            sys.exit(1)

    headers = {
        'apikey': service_key,
        'Authorization': f'Bearer {service_key}',
    }

    import openpyxl
    wb = openpyxl.load_workbook(args.xlsx_path, read_only=True, data_only=True)

    if 'Outreach Tracker' in wb.sheetnames:
        ws = wb['Outreach Tracker']
    else:
        ws = wb.active

    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Row index 0 = title, 1 = column headers, 2+ = data
    raw_headers = [cell_str(v) for v in all_rows[1]]

    rows = []
    row_number = 0
    for raw_row in all_rows[2:]:
        xlsx_row = {raw_headers[i]: raw_row[i] for i in range(len(raw_headers))}
        org_name = cell_str(xlsx_row.get('Organization Name'))
        if not org_name:
            continue
        row_number += 1
        record = map_row(xlsx_row, args.chapter_id)
        record['row_number'] = row_number
        rows.append(record)

    total    = len(rows)
    errors   = 0
    inserted = 0

    print(f"Found {total} rows to migrate (chapter_id={args.chapter_id})")
    if args.dry_run:
        print('[dry-run mode — no data will be inserted]')
        print('\nFirst 3 rows (order check):')
        for r in rows[:3]:
            print(f"  row_number={r['row_number']}  org_name={r['org_name']}")
        print()

    for i in range(0, total, BATCH_SIZE):
        batch     = rows[i:i + BATCH_SIZE]
        batch_end = min(i + BATCH_SIZE, total)
        errs      = insert_batch(batch, supabase_url, headers, args.dry_run)
        errors   += errs
        inserted += len(batch) - errs
        if not args.dry_run:
            print(f"Inserted row {batch_end}/{total}...")

    print(f"\nMigration complete. {inserted} rows inserted, {errors} errors.")


if __name__ == '__main__':
    main()
